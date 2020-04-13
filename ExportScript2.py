import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
import win32com.client as win32
from win32com.client import Dispatch
import TestCommander
import os
import sys
from PyQt5 import QtCore, QtGui, QtWidgets
from datetime import datetime
from distutils.dir_util import copy_tree

excel = win32.gencache.EnsureDispatch('Excel.Application')
excel.Visible = False
today = str(datetime.today().strftime('%d-%m-%Y'))

def createFolder(directory):
    try:
        if not os.path.exists(directory):
            os.mkdir(directory)
    except OSError:
        print('Error: Creating directory. ' + directory)

class TASISCopyResults(QtWidgets.QMainWindow, TestCommander.Ui_Exporter):

    def __init__(self, parent = None):
        super().__init__(parent)
        self.setupUi(self)
        self.show()
        self.cancelButton.clicked.connect(self.functionCancel)
        self.refreshButton.clicked.connect(self.functionRefresh)
        self.copyButton.clicked.connect(self.functionCopy)
        self.excelFile.clear()

    def functionRefresh(self):
        self.excelFile.clear()
        self.excelFile.addItem(excel.ActiveWorkbook.FullName)
        self.progressBar.setValue(0)
        self.copyButton.setEnabled(True)

    def functionCopy(self):

        path_BaseWb = (str(excel.ActiveWorkbook.FullName))
        path = path_BaseWb
        print(path)
        # New Folder containing test results creation
        folderName = str(excel.ActiveWorkbook.Name)
        folderName = folderName.split(".xlsx")[0]
        newWb_filename = folderName + '_new' + '.xlsx'
        folderName = folderName + "_" + today
        print(folderName)

        path = str(path.split(excel.ActiveWorkbook.Name)[0] + folderName + "\\")
        print('Root path in which we put results: ' + path)
        createFolder(path)

        path_NewWb = path + newWb_filename
        print(path_NewWb)
        newWB = Workbook()
        newWB.save(filename=path_NewWb)
        baseWb = excel.Workbooks.Open(Filename=path_BaseWb)
        newWB = excel.Workbooks.Open(Filename=path_NewWb)
        baseWorksheet = baseWb.Worksheets(1)
        baseWorksheet.Copy(Before=newWB.Worksheets(1))
        newWB.Close(SaveChanges=True)
        baseWb.Close(SaveChanges=False)
        #newWB = excel.Workbooks.Open(Filename=path_NewWb)
        #excel.Visible = False
        workplm = load_workbook(path_NewWb)
        firstSheet = workplm.active
        resultCol = firstSheet['P']
        firstCol = firstSheet['A']
        print(len(resultCol))
        self.progressBar.setMaximum(len(resultCol)-1)
        hyperLink = '=HYPERLINK('
        for x in range(len(firstCol)):
            if (firstCol[x].value == None) or (x == 0) or (resultCol[x].value == None):
                self.progressBar.setValue(x)
                print('BLANK')
            else:
                testID_DIR = str(firstCol[x].value)
                reportDir = str(resultCol[x].value)
                aux = reportDir
                aux = aux.split(",")[0]
                testResult = reportDir.split(',')[1]
                aux = aux.strip(hyperLink)
                aux = aux.strip('"')
                print(aux)
                finalPath = (path + testID_DIR)
                print(resultCol[x].value)
                resultCol[x].value = str(hyperLink + '"' + finalPath + '"' + ',' + testResult)
                print(resultCol[x].value)
                createFolder(finalPath)
                copy_tree(aux, finalPath)
                self.progressBar.setValue(x)
                workplm.save(path_NewWb)


        excel.Visible = True
        #newWB.Close(SaveChanges=True)
        excel.Quit()

    def functionCancel(self):
        excel.Visible = True
        self.close()

def main():
    app = QtWidgets.QApplication(sys.argv)
    main = TASISCopyResults()
    app.exec_()

if __name__ == '__main__':
    main()
